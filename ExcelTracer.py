# ExcelTracer acts as a go-between for debugging Excel-/Spreadsheet-based models. 
# I got VERY tired of switching between sheets looking for named cell ranges and which formulas they are referenced in. 
#
# End goal would be to more effectively be able to trace dependents and trace prededents. Tracking where cells are used 
# will mitigate any blowups or #REF outs, at least it should make them significantly less likely. 
#
#
#
#######################################################################################################################

import openpyxl as xl
from openpyxl.utils import get_column_letter
import re
import os


def read_spreadsheet(fpath, fname, sheet_name=None, verbose=False):

    full_path = os.path.join(fpath, fname)
    if verbose == True:
        print(f'Reading {fname} from path {fpath})

        if sheet_name == None:
            print('Searching all sheets...')

        else:
            print(f'Searching sheet {sheet_name}...')

    try:
        wb = xl.load_workbook(full_path)

    except Exception as err:
        print(err)

    if sheet_name == None:
        try:
            full_book = {sheet : wb[sheet] for sheet in wb.sheetnames}
            return full_book

        except Exception as err:
            print(err)

    else:
        return {sheet_name : wb[sheet_name]}


def search_sheet(sheet, to_search=None, verbose=False):

    if to_search != None:
        reg = re.compile(to_search, re.IGNORECASE)

    else: raise ValueError('Must enter a search value - please try again.')

    found = list()

    for row in sheet.iter_rows():
        for cell in row:
            val = cell.value
            if val != None:
                match = reg.search(str(val))
                if match:
                    found_cell = get_column_letter(cell.column) + str(cell.row)

                    if verbose == True:
                        print(f'Found searched value in cell {found_cell}')

                    found.append(found_cell)


    if verbose == True:
        if len(found) == 0:
            print(f'{sheet.title} does not contain searched string "{to_search}"')

    return found


def search_multiple(wb_dict, sheet_names=None, search_term=None):
    found = dict()
    if sheet_names == None:
        for ws_name, ws in wb_dict.items():
            found_list = search_sheet(ws, search_term)
            if len(found_list) == 0:
                continue
            else:
                found[ws_name] = found_list

        return found


    else:
        for sheet_name in sheet_names:
            found_lst = search_sheet(wb_dict[sheet_name], search_term)
            if len(found_lst) == 0:
                continue

            else:
                found[sheet_name] = found_lst

        return found
